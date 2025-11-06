"""
Utilitaires pour l'application
"""
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from django.core.cache import cache
from django.utils import timezone
from datetime import timedelta
from .models import ActionLog


def log_action(
    user, action_type, model_name, description, object_id=None, request=None
):
    """Enregistrer une action dans les logs"""
    ip_address = None
    user_agent = ""

    if request:
        ip_address = get_client_ip(request)
        user_agent = request.META.get("HTTP_USER_AGENT", "")

    ActionLog.objects.create(
        user=user,
        action_type=action_type,
        model_name=model_name,
        object_id=object_id,
        description=description,
        ip_address=ip_address,
        user_agent=user_agent,
    )


def get_client_ip(request):
    """Récupérer l'adresse IP du client"""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0]
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


def get_cached_stats(key, compute_func, timeout=300):
    """Récupérer des statistiques depuis le cache ou les calculer"""
    stats = cache.get(key)
    if stats is None:
        stats = compute_func()
        cache.set(key, stats, timeout)
    return stats


def export_to_excel(queryset, filename, headers, data_func):
    """Exporter un queryset vers Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    # Headers
    ws.append(headers)

    # Data
    for obj in queryset:
        row = data_func(obj)
        ws.append(row)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'

    wb.save(response)
    return response


def import_from_excel(file, model_class, mapping_func, error_handler=None):
    """Importer des données depuis un fichier Excel"""
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    imported = 0
    errors = []

    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = mapping_func(row)
            if data:
                obj = model_class(**data)
                obj.save()
                imported += 1
        except Exception as e:
            error_msg = f"Row {row_idx}: {str(e)}"
            errors.append(error_msg)
            if error_handler:
                error_handler(row_idx, e)

    return imported, errors


def validate_file_upload(
    file, allowed_extensions=[".xlsx", ".xls", ".csv"], max_size=5242880
):
    """Valider un fichier uploadé"""
    import os

    # Vérifier l'extension
    ext = os.path.splitext(file.name)[1].lower()
    if ext not in allowed_extensions:
        return (
            False,
            f"Extension non autorisée. Extensions autorisées: {', '.join(allowed_extensions)}",
        )

    # Vérifier la taille
    if file.size > max_size:
        return (
            False,
            f"Fichier trop volumineux. Taille maximale: {max_size / 1024 / 1024}MB",
        )

    return True, None
