"""
Utilitaires pour la génération de rapports PDF et Excel
"""
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal


def generate_pdf_report(title, data, filename):
    """Génère un rapport PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=18,
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#1e40af"),
        spaceAfter=30,
        alignment=TA_CENTER,
    )

    # Contenu
    story = []
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Tableau de données
    if data:
        table_data = []
        # En-têtes
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            table_data.append([Paragraph(str(h), styles["Heading3"]) for h in headers])
            # Données
            for row in data:
                table_data.append([str(row.get(h, "")) for h in headers])
        else:
            table_data = data

        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.lightgrey],
                    ),
                ]
            )
        )

        story.append(table)

    # Date de génération
    story.append(Spacer(1, 0.2 * inch))
    story.append(
        Paragraph(
            f"<i>Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}</i>",
            styles["Normal"],
        )
    )

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def generate_excel_report(title, data, filename):
    """Génère un rapport Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport"

    # Style pour les en-têtes
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    # Titre
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Données
    if data:
        row_num = 3
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
            # En-têtes
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Données
            for row_data in data:
                row_num += 1
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    value = row_data.get(header, "")
                    cell.value = str(value) if value is not None else ""
        else:
            # Données sous forme de liste de listes
            for row_data in data:
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    if row_num == 3:  # En-têtes
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.value = str(value) if value is not None else ""
                row_num += 1

        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    # Date de génération
    ws.cell(
        row=row_num + 2, column=1
    ).value = f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"

    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
